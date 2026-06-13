import openpyxl
from datetime import datetime

ARCHIVO_EXCEL = "../datos/Base_Datos_Soporte.xlsx"

ESTADO_INICIO = "INICIO"
ESTADO_VALIDANDO_USUARIO = "VALIDANDO_USUARIO"
ESTADO_ESPERANDO_PROBLEMA = "ESPERANDO_PROBLEMA"
ESTADO_BUSCANDO_FAQ = "BUSCANDO_FAQ"
ESTADO_ESPERANDO_CONFIRMACION = "ESPERANDO_CONFIRMACION"
ESTADO_GENERANDO_TICKET = "GENERANDO_TICKET"
ESTADO_DERIVADO_N2 = "DERIVADO_N2"
ESTADO_CERRADO = "CERRADO"


def abrir_bd():
    return openpyxl.load_workbook(ARCHIVO_EXCEL)


def buscar_usuario(legajo):

    wb = abrir_bd()
    hoja = wb["Usuarios_Empleados"]

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        if str(fila[0]) == str(legajo):

            return {
                "legajo": fila[0],
                "nombre": fila[1],
                "sector": fila[2],
                "email": fila[3]
            }

    return None


def buscar_solucion(problema):

    wb = abrir_bd()
    hoja = wb["Base_Conocimientos_FAQ"]

    problema = problema.lower()

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        palabras_clave = str(fila[2]).lower()

        lista_palabras = palabras_clave.split(",")

        for palabra in lista_palabras:

            palabra = palabra.strip()

            if palabra in problema:

                return fila[3]

    return None


def calcular_prioridad(problema):

    problema = problema.lower()

    alta = [
        "servidor",
        "base de datos",
        "caido",
        "caído"
    ]

    media = [
        "internet",
        "vpn",
        "red"
    ]

    for palabra in alta:

        if palabra in problema:
            return "ALTA"

    for palabra in media:

        if palabra in problema:
            return "MEDIA"

    return "BAJA"


def generar_ticket(legajo, problema, prioridad):

    wb = abrir_bd()
    hoja = wb["Registro_Tickets"]

    ultimo_id = 0

    for fila in hoja.iter_rows(min_row=2, values_only=True):

        if fila[0]:

            try:
                ultimo_id = max(
                    ultimo_id,
                    int(fila[0])
                )
            except:
                pass

    nuevo_id = ultimo_id + 1

    hoja.append([
        nuevo_id,
        legajo,
        datetime.now().strftime("%d/%m/%Y %H:%M"),
        problema,
        prioridad,
        "ABIERTO",
        ""
    ])

    wb.save(ARCHIVO_EXCEL)

    return nuevo_id


def main():

    print("\n==============================")
    print("MESA DE AYUDA IT")
    print("==============================")

    estado = ESTADO_VALIDANDO_USUARIO

    legajo = input("\nIngrese su legajo: ")

    if not legajo.isdigit():

        print("Error: debe ingresar un número.")
        return

    usuario = buscar_usuario(legajo)

    if usuario is None:

        print("Legajo inexistente.")
        return

    print(f"\nBienvenido {usuario['nombre']}")

    estado = ESTADO_ESPERANDO_PROBLEMA

    problema = input(
        "\nDescriba su problema: "
    ).strip()

    if problema == "":

        print("Debe ingresar un problema.")
        return

    estado = ESTADO_BUSCANDO_FAQ

    solucion = buscar_solucion(problema)

    if solucion:

        print("\nSolución encontrada:")
        print(solucion)

        estado = ESTADO_ESPERANDO_CONFIRMACION

        respuesta = input(
            "\n¿Se resolvió el problema? (SI/NO): "
        ).upper()

        while respuesta not in ["SI", "NO"]:

            respuesta = input(
                "Ingrese SI o NO: "
            ).upper()

        if respuesta == "SI":

            estado = ESTADO_CERRADO

            print("\nIncidente resuelto.")
            print("Estado:", estado)

            return

    estado = ESTADO_GENERANDO_TICKET

    prioridad = calcular_prioridad(problema)

    ticket = generar_ticket(
        legajo,
        problema,
        prioridad
    )

    print(f"\nTicket generado N° {ticket}")
    print("Prioridad:", prioridad)

    if prioridad == "ALTA":

        estado = ESTADO_DERIVADO_N2

        print("Derivado a Soporte Nivel 2.")
        print("Estado:", estado)

    else:

        estado = ESTADO_CERRADO

        print("Será atendido por Mesa de Ayuda.")
        print("Estado:", estado)


if __name__ == "__main__":
    main()
